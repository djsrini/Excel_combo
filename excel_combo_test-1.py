import openpyxl
import os
import glob
from openpyxl import Workbook

loc = os.getcwd()
print(loc)
fileDir = glob.glob( str(loc) +"\*.xlsx")
fileNam = []
fileName = []
array = []
print(fileDir)

writeBook = Workbook()
sheetWrite= writeBook.active

count = 0

for i in fileDir:
    if loc in i:
        temp = i.replace(loc, '')
        fileNam.append(temp)

for i in fileNam:
    temp = i
    temp = temp[1:]
    fileName.append(temp)


print(fileName[1])

row = 0
column = 0
sheet = ""
temp = 0

for i in fileName:
    book = openpyxl.load_workbook(i)
    sheet = book.active
    # count = 0
    # rowvalue = 0
    # columnvalue = 0

    rowT = sheet.max_row
    columnT = sheet.max_column

    if columnT > column:
        column = columnT

    row = row + rowT

    print(str(rowT) + " " + str(columnT))
    print(str(row) + " " + str(column))


    # print(str(sheet.cell(column= 1, row=1).value))

    a = [[0] * column for e in range(row)]

    for j in range(rowT):
        for k in range(columnT):
            a[j][k] = sheet.cell(column=k + 1, row=j + 1).value
        sheetWrite.append(a[j])

    # for j in range(row):
    #     for k in range(column):
    #         a[j][k] = sheet.cell(column=k + 1, row=j + 1).value
    #         sheetWrite.cell(row=temp, column=k + 1).value = str(a[j][k])

    print(a)
    print("--------------------------------------------------------")

print(a)
writeBook.save("Book2.xlsx")